# Copyright (c) 2026, Oliver Reid and contributors
# For license information, please see license.txt

"""
API endpoints for NCE Sync app.
"""

import frappe
from frappe import _


@frappe.whitelist()
def toggle_auto_sync(table_names):
	"""
	Toggle the auto_sync_active field for the specified WP Tables.

	Args:
		table_names: List of WP Tables names (or JSON string)

	Returns:
		str: Summary message of changes made
	"""
	if isinstance(table_names, str):
		import json

		table_names = json.loads(table_names)

	if not table_names:
		frappe.throw(_("No tables specified"))

	enabled_count = 0
	disabled_count = 0

	for table_name in table_names:
		doc = frappe.get_doc("WP Tables", table_name)

		if doc.auto_sync_active:
			doc.auto_sync_active = 0
			disabled_count += 1
		else:
			doc.auto_sync_active = 1
			enabled_count += 1

		doc.save(ignore_permissions=True)

	frappe.db.commit()

	# Build response message
	parts = []
	if enabled_count:
		parts.append(f"{enabled_count} enabled")
	if disabled_count:
		parts.append(f"{disabled_count} disabled")

	return _("Auto sync: {0}").format(", ".join(parts))


@frappe.whitelist()
def get_table_links_grid_data():
	"""
	Return mirrored tables and their Link field relationships for the grid UI.
	Data is derived from actual DocType metas (no stored copy).

	Returns:
		dict: {
			"tables": [{"doctype": "Events", "label": "Events"}, ...],
			"links": {"Events": {"Venues": [{"field": "venue_id", "label": "Venue"}]}}
		}
	"""
	tables = frappe.get_all(
		"WP Tables",
		filters={"mirror_status": ["in", ["Mirrored", "Linked"]], "frappe_doctype": ["!=", ""]},
		fields=["frappe_doctype", "nce_name", "table_name"],
		order_by="frappe_doctype",
	)

	# Build list of mirrored DocTypes with display labels
	doctypes = []
	seen = set()
	for row in tables:
		dt = row.get("frappe_doctype")
		if not dt or dt in seen:
			continue
		seen.add(dt)
		label = row.get("nce_name") or row.get("table_name") or dt
		doctypes.append({"doctype": dt, "label": label})

	# Scan each DocType's meta for Link fields pointing to other mirrored tables
	mirrored_set = {d["doctype"] for d in doctypes}
	links = {}  # source_doctype -> { target_doctype -> [{"field", "label"}, ...] }

	for d in doctypes:
		source = d["doctype"]
		try:
			meta = frappe.get_meta(source)
		except Exception:
			continue

		for df in meta.fields:
			if df.fieldtype != "Link" or not df.options:
				continue
			target = df.options
			if target not in mirrored_set or target == source:
				continue

			if source not in links:
				links[source] = {}
			if target not in links[source]:
				links[source][target] = []
			links[source][target].append(
				{
					"field": df.fieldname,
					"label": df.label or df.fieldname,
					"many_doctype": source,
				}
			)

			# Also record in the reverse direction so the grid cell works both ways
			if target not in links:
				links[target] = {}
			if source not in links[target]:
				links[target][source] = []
			links[target][source].append(
				{
					"field": df.fieldname,
					"label": df.label or df.fieldname,
					"many_doctype": source,
				}
			)

	return {"tables": doctypes, "links": links}


@frappe.whitelist()
def apply_table_link_changes(to_add, to_delete):
	"""
	Apply pending link field changes in batch.

	Args:
		to_add: JSON list of {"many_doctype", "one_doctype", "field_name"}
		to_delete: JSON list of {"many_doctype", "field_name"}

	Returns:
		str: Summary message
	"""
	import json as _json

	additions = _json.loads(to_add) if isinstance(to_add, str) else to_add
	deletions = _json.loads(to_delete) if isinstance(to_delete, str) else to_delete
	msgs = []

	for item in deletions:
		dt = item["many_doctype"]
		fname = item["field_name"]
		frappe.clear_cache(doctype=dt)
		meta = frappe.get_meta(dt)
		existing = meta.get_field(fname) if meta.has_field(fname) else None
		if not existing or existing.fieldtype != "Link":
			msgs.append(_("{0}.{1} is not a Link field, skipped").format(dt, fname))
			continue
		try:
			doc = frappe.get_doc("DocType", dt)
			for f in doc.fields:
				if f.fieldname == fname:
					f.fieldtype = "Data"
					f.options = ""
					break
			doc.save(ignore_permissions=True)
			frappe.db.commit()
			frappe.clear_cache(doctype=dt)
			msgs.append(_("Reverted {0}.{1} from Link to Data").format(dt, fname))
		except Exception as e:
			msgs.append(_("FAILED revert {0}.{1}: {2}").format(dt, fname, str(e)))

	for item in additions:
		dt = item["many_doctype"]
		one_dt = item["one_doctype"]
		fname = item["field_name"]
		frappe.clear_cache(doctype=dt)
		meta = frappe.get_meta(dt)
		existing = meta.get_field(fname) if meta.has_field(fname) else None

		if existing and existing.fieldtype == "Link" and existing.options == one_dt:
			msgs.append(_("{0}.{1} already links to {2}, skipped").format(dt, fname, one_dt))
			continue

		try:
			doc = frappe.get_doc("DocType", dt)
			if existing:
				for f in doc.fields:
					if f.fieldname == fname:
						f.fieldtype = "Link"
						f.options = one_dt
						break
				action = _("Converted {0}.{1} ({2}) to Link → {3}").format(
					dt, fname, existing.fieldtype, one_dt
				)
			else:
				doc.append(
					"fields",
					{
						"fieldname": fname,
						"fieldtype": "Link",
						"label": fname.replace("_", " ").title(),
						"options": one_dt,
					},
				)
				action = _("Added {0}.{1} → {2}").format(dt, fname, one_dt)
			doc.save(ignore_permissions=True)
			frappe.db.commit()
			frappe.clear_cache(doctype=dt)
			msgs.append(action)
		except Exception as e:
			msgs.append(_("FAILED {0}.{1}: {2}").format(dt, fname, str(e)))

	return "; ".join(msgs) if msgs else _("No changes applied")


@frappe.whitelist()
def export_all_to_excel(doctype):
	"""Enqueue a background job that exports the full DocType to xlsx."""
	if not frappe.has_permission(doctype, "read"):
		frappe.throw(_("Not permitted"), frappe.PermissionError)

	total = frappe.db.count(doctype)

	frappe.enqueue(
		_build_excel_file,
		doctype=doctype,
		user=frappe.session.user,
		queue="default",
		is_async=True,
	)

	return total


@frappe.whitelist()
def sync_doctype_rows(doctype, names):
	"""
	Queue a job to fetch listed rows from WordPress and upsert into Frappe.

	Runs on the ``default`` queue (serialized with scheduled table sync).

	Args:
		doctype (str): Frappe DocType name.
		names (list | str): List of ``name`` values or JSON-encoded list.

	Returns:
		dict: ``queued``, ``doctype``, ``row_count``, and a short ``message``.
	"""
	import json

	if not frappe.has_permission(doctype, "read"):
		frappe.throw(_("Not permitted"), frappe.PermissionError)

	if isinstance(names, str):
		names = json.loads(names)

	if not isinstance(names, (list, tuple)):
		frappe.throw(_("names must be a list"))

	if not names:
		frappe.throw(_("No rows specified"))

	raw_names = [n for n in names if n is not None]
	if not raw_names:
		frappe.throw(_("No row names supplied"))

	wp_table_matches = frappe.get_all(
		"WP Tables",
		filters={
			"frappe_doctype": doctype,
			"mirror_status": ["in", ["Mirrored", "Linked"]],
		},
		limit_page_length=1,
		pluck="name",
	)
	if not wp_table_matches:
		frappe.throw(
			_("No WP Tables configuration found for DocType '{0}' with status Mirrored or Linked").format(
				doctype
			)
		)

	wp_table_doc = frappe.get_doc("WP Tables", wp_table_matches[0])
	if getattr(wp_table_doc, "doctype_source", "") == "Native" or not wp_table_doc.table_name:
		frappe.throw(_("DocType '{0}' is not linked to a WordPress table").format(doctype))

	wp_conn = frappe.get_single("WordPress Connection")
	if not wp_conn.host:
		frappe.throw(_("WordPress Connection not configured"))

	frappe.enqueue(
		"nce_sync.utils.data_sync.run_sync_doctype_rows_job",
		queue="default",
		timeout=3600,
		doctype=doctype,
		names=list(raw_names),
		user=frappe.session.user,
	)

	frappe.msgprint(
		_(
			"Row sync queued for {0} ({1} name(s)); it runs on the same queue as table sync "
			"and you will get a toast when it finishes."
		).format(doctype, len(raw_names)),
		indicator="blue",
		alert=True,
	)

	return {
		"queued": True,
		"doctype": doctype,
		"row_count": len(raw_names),
		"message": _("Queued on default worker queue"),
	}


@frappe.whitelist()
def sync_linked_doctype_rows(doctype, link_field, link_value):
	"""
	Queue a job: delete existing Frappe rows for a Link filter, then re-insert from WordPress.

	Frappe rows where ``link_field`` = ``link_value`` are removed. WordPress rows whose
	mapped column for ``link_field`` equals ``link_value`` are then pulled and upserted.

	Requires a **Link** DocField on ``doctype`` with a matching column in WP Tables mapping.
	"""
	from frappe.utils import cstr

	from nce_sync.utils.column_mapper import build_reverse_mapping, load_column_mapping

	if not frappe.has_permission(doctype, "read"):
		frappe.throw(_("Not permitted"), frappe.PermissionError)

	if not frappe.has_permission(doctype, "delete"):
		frappe.throw(_("Delete permission required on {0}").format(doctype), frappe.PermissionError)

	link_field = (link_field or "").strip()
	link_value = cstr(link_value).strip() if link_value is not None else ""
	if not link_field:
		frappe.throw(_("link_field is required"))

	if not link_value:
		frappe.throw(_("link_value is required"))

	meta = frappe.get_meta(doctype)
	df = meta.get_field(link_field)
	if not df:
		frappe.throw(_("Field '{0}' not found on DocType '{1}'").format(link_field, doctype))
	if df.fieldtype != "Link":
		frappe.throw(_("Field '{0}' must be a Link field (got {1})").format(link_field, df.fieldtype))

	wp_table_matches = frappe.get_all(
		"WP Tables",
		filters={
			"frappe_doctype": doctype,
			"mirror_status": ["in", ["Mirrored", "Linked"]],
		},
		limit_page_length=1,
		pluck="name",
	)
	if not wp_table_matches:
		frappe.throw(
			_("No WP Tables configuration found for DocType '{0}' with status Mirrored or Linked").format(
				doctype
			)
		)

	wp_table_doc = frappe.get_doc("WP Tables", wp_table_matches[0])
	if getattr(wp_table_doc, "doctype_source", "") == "Native" or not wp_table_doc.table_name:
		frappe.throw(_("DocType '{0}' is not linked to a WordPress table").format(doctype))

	wp_conn = frappe.get_single("WordPress Connection")
	if not wp_conn.host:
		frappe.throw(_("WordPress Connection not configured"))

	reverse_mapping = build_reverse_mapping(load_column_mapping(wp_table_doc))
	if not reverse_mapping.get(link_field):
		frappe.throw(
			_("No WordPress column mapped for Link field '{0}' on DocType '{1}'").format(
				link_field, doctype
			)
		)

	frappe.enqueue(
		"nce_sync.utils.data_sync.run_sync_linked_doctype_rows_job",
		queue="default",
		timeout=3600,
		doctype=doctype,
		link_field=link_field,
		link_value=link_value,
		user=frappe.session.user,
	)

	frappe.msgprint(
		_(
			"Linked row sync queued for {0}: {1} = {2}. Rows will be rebuilt from WordPress."
		).format(doctype, link_field, link_value),
		indicator="blue",
		alert=True,
	)

	return {
		"queued": True,
		"doctype": doctype,
		"link_field": link_field,
		"link_value": link_value,
		"message": _("Queued on default worker queue"),
	}


def _build_excel_file(doctype, user):
	"""Background job: build xlsx and send download URL via realtime."""
	import io

	from openpyxl import Workbook
	from openpyxl.utils import get_column_letter

	meta = frappe.get_meta(doctype)
	skip_types = frozenset(
		{
			"Section Break",
			"Column Break",
			"Tab Break",
			"HTML",
			"Fold",
			"Heading",
		}
	)
	fields = [df for df in meta.fields if df.fieldtype not in skip_types]
	fieldnames = ["name"] + [df.fieldname for df in fields]
	labels = ["ID"] + [df.label or df.fieldname for df in fields]

	rows = frappe.get_all(
		doctype,
		fields=fieldnames,
		limit_page_length=0,
		order_by="name asc",
	)

	wb = Workbook()
	ws = wb.active
	ws.title = doctype[:31]

	ws.append(labels)
	for cell in ws[1]:
		cell.font = cell.font.copy(bold=True)

	for row in rows:
		ws.append([row.get(f) for f in fieldnames])

	for idx, _ in enumerate(labels, 1):
		ws.column_dimensions[get_column_letter(idx)].width = 18

	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)

	fname = f"{doctype.replace(' ', '_')}.xlsx"
	file_doc = frappe.get_doc(
		{
			"doctype": "File",
			"file_name": fname,
			"content": buf.getvalue(),
			"is_private": 1,
		}
	)
	file_doc.save(ignore_permissions=True)
	frappe.db.commit()

	frappe.publish_realtime(
		"excel_export_ready",
		{"file_url": file_doc.file_url},
		user=user,
	)
